
from pathlib import Path     
from collections import deque 
import numpy as np            
import pandas as pd           
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PASTA = Path(__file__).resolve().parent

ARQ_COTACOES = PASTA / 'COTAÇÕES - CASE.xlsx'
ARQ_SALESFORCE = PASTA / 'SalesForce - CASE.xlsx'
ARQ_SAIDA = PASTA / 'case_b2b_resultado.xlsx'

JANELA_DIAS = 30
TIPOS_VALIDOS = {'Banda', 'PrecoFixo'}
STATUS_APROVADO = 'Aprovada'


def verificar_arquivos() -> None: 
    faltando = [str(p) for p in (ARQ_COTACOES, ARQ_SALESFORCE) if not p.exists()]
    if faltando:
        raise FileNotFoundError(
            'Não encontrei o(s) arquivo(s) abaixo. Coloque-os na mesma pasta deste script:\n'
            + '\n'.join(faltando)
        )


def normalizar_chave(serie: pd.Series) -> pd.Series: 
    s = serie.astype('string').str.strip().str.replace(r'\.0$', '', regex=True)
    return s.str.lstrip('0').replace('', pd.NA)


def ler_bases() -> tuple[pd.DataFrame, pd.DataFrame]:
    verificar_arquivos()
    cot = pd.read_excel(ARQ_COTACOES, sheet_name='Base', header=2,
                        engine='openpyxl', dtype=str)
    sf = pd.read_excel(ARQ_SALESFORCE, sheet_name='Base 2',
                       engine='openpyxl', dtype=str)
    return cot, sf


def preparar_cotacoes(df: pd.DataFrame) -> pd.DataFrame:
    x = df.copy()
    x['_ordem_original'] = np.arange(len(x))
    x['Dt. Criação Sol.'] = pd.to_datetime(x['Dt. Criação Sol.'], dayfirst=True, errors='coerce')
    x['Última Ação em'] = pd.to_datetime(x['Última Ação em'], errors='coerce')
    for col in ['Preço Atual', 'Preço Proposto', 'Var. Preço Proposto']:
        x[col] = pd.to_numeric(x[col], errors='coerce')
    x['_id_sap'] = normalizar_chave(x['ID da Cotação do SAP'])
    x['_cliente'] = normalizar_chave(x['Código do Cliente'])
    x['_material'] = normalizar_chave(x['Material'])
    x['_expedicao'] = normalizar_chave(x['Cód. Expedição'])
    return x


def preparar_salesforce(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    x = df.copy()
    x['_id_sap'] = normalizar_chave(x['ID da Cotação do SAP'])
    x['_material'] = normalizar_chave(x['Material: Código do material'])
    x['_expedicao'] = normalizar_chave(x['Cód. Expedição'])
    x['Frete Comercial_num'] = pd.to_numeric(x['Frete Comercial'], errors='coerce')
    x['Chapa_num'] = pd.to_numeric(x['Chapa'], errors='coerce')
    chave = ['_id_sap', '_material', '_expedicao']

    conflitos = (x.groupby(chave, dropna=False)
                   .agg(Linhas=('TipoCotação', 'size'),
                        Valores_Frete=('Frete Comercial_num', lambda s: s.dropna().nunique()),
                        Valores_Chapa=('Chapa_num', lambda s: s.dropna().nunique()))
                   .reset_index())
    conflitos = conflitos[(conflitos['Linhas'] > 1) |
                          (conflitos['Valores_Frete'] > 1) |
                          (conflitos['Valores_Chapa'] > 1)]

    mapa = (x.sort_index()
             .groupby(chave, dropna=False, as_index=False)
             .agg(**{'Frete Comercial': ('Frete Comercial_num', 'first'),
                     'Chapa': ('Chapa_num', 'first')}))
    return mapa, conflitos


def adicionar_salesforce(cot: pd.DataFrame, mapa: pd.DataFrame) -> pd.DataFrame:
    x = cot.merge(mapa, on=['_id_sap', '_material', '_expedicao'], how='left', validate='m:1')
    x['Match SalesForce'] = np.where(x['Frete Comercial'].notna() | x['Chapa'].notna(), 'Sim', 'Não')
    return x


def calcular_historico_30d(df: pd.DataFrame) -> pd.DataFrame:
    x = df.copy()
    x['Qtd. Aprovadas Anteriores 30d'] = 0
    x['Soma Var. Aprovada Anterior 30d'] = 0.0
    chave = ['_cliente', '_material', '_expedicao']

    for _, idx in x.groupby(chave, dropna=False).groups.items():
        ordem = x.loc[idx].sort_values(['Dt. Criação Sol.', 'Última Ação em', '_ordem_original'],
                                       na_position='last').index.tolist()
        janela = deque()  # cada item guardado é (data, id_documento, variacao)
        docs_na_janela = set()
        pos = 0
        while pos < len(ordem):
            data_atual = x.at[ordem[pos], 'Dt. Criação Sol.']
            lote = []
            while pos < len(ordem) and x.at[ordem[pos], 'Dt. Criação Sol.'] == data_atual:
                lote.append(ordem[pos]); pos += 1

            if pd.notna(data_atual):
                limite = data_atual - pd.Timedelta(days=JANELA_DIAS)
                while janela and janela[0][0] < limite:
                    _, doc_antigo, _ = janela.popleft()
                    docs_na_janela.discard(doc_antigo)
                qtd = len(janela)
                soma = sum(item[2] for item in janela)
                x.loc[lote, 'Qtd. Aprovadas Anteriores 30d'] = qtd
                x.loc[lote, 'Soma Var. Aprovada Anterior 30d'] = soma

                # Só depois de calcular o lote é que suas próprias aprovações
                # (se houver) entram na janela, para valerem em linhas futuras.
                for i in lote:
                    elegivel = (x.at[i, 'Status da Cotação'] == STATUS_APROVADO and
                                x.at[i, 'TipoCotação'] in TIPOS_VALIDOS)
                    doc = x.at[i, '_id_sap']
                    if elegivel and pd.notna(doc) and doc not in docs_na_janela:
                        variacao = x.at[i, 'Var. Preço Proposto']
                        variacao = 0.0 if pd.isna(variacao) else float(variacao)
                        janela.append((data_atual, doc, variacao))
                        docs_na_janela.add(doc)
    return x


def criar_tabelas_dashboard(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    data_ref = df['Dt. Criação Sol.'].max()
    inicio = data_ref - pd.Timedelta(days=JANELA_DIAS)
    aprov_30 = df[(df['Dt. Criação Sol.'].between(inicio, data_ref)) &
                  (df['Status da Cotação'] == STATUS_APROVADO)].copy()
    aprov_30['Desconto Concedido'] = (-aprov_30['Var. Preço Proposto'].fillna(0)).clip(lower=0)

    top_clientes = (aprov_30.groupby('Código do Cliente', dropna=False)['Desconto Concedido']
                    .sum().sort_values(ascending=False).head(10).reset_index())
    top_regiao = (aprov_30.groupby(['Código do Cliente', 'Região', 'Gestão de Vendas'], dropna=False)
                  ['Desconto Concedido'].sum().sort_values(ascending=False).head(10).reset_index())

    cond = (aprov_30['Cliente Condição'].fillna('Sem informação').value_counts()
            .rename_axis('Cliente Condição').reset_index(name='Qtd. Cotações'))
    total_cond = cond['Qtd. Cotações'].sum()
    cond['Representatividade'] = cond['Qtd. Cotações'] / total_cond if total_cond else 0

    cif = df[df['Incoterm'].eq('CIF')].copy()
    cif['Frete Zerado'] = pd.to_numeric(cif['Frete Comercial'], errors='coerce').eq(0)
    frete = (cif.groupby('Cód. Expedição', dropna=False)
             .agg(**{'Qtd. CIF': ('Incoterm', 'size'),
                     'Qtd. Frete Zero': ('Frete Zerado', 'sum')}).reset_index())
    frete['Representatividade Frete Zero'] = frete['Qtd. Frete Zero'] / frete['Qtd. CIF']
    frete = frete.sort_values(['Qtd. Frete Zero', 'Qtd. CIF'], ascending=False)

    parametros = pd.DataFrame({
        'Parâmetro': ['Data de referência', 'Início da janela inclusiva', 'Janela (dias)',
                      'Tipos válidos no histórico', 'Status aprovado',
                      'Chave do merge', 'Chave do histórico'],
        'Valor': [data_ref, inicio, JANELA_DIAS, 'Banda e PrecoFixo', STATUS_APROVADO,
                  'ID SAP + Material + Expedição', 'Cliente + Material + Expedição']})
    return {'Top_10_Clientes': top_clientes, 'Top_10_Regiao_Gestao': top_regiao,
            'Resumo_Cliente_Condicao': cond, 'Resumo_Frete_CIF': frete,
            'Parametros': parametros}


def exportar_excel(df: pd.DataFrame, dashboards: dict[str, pd.DataFrame],
                   conflitos: pd.DataFrame) -> None:
    tecnicas = [c for c in df.columns if c.startswith('_')]
    base_saida = df.sort_values('_ordem_original').drop(columns=tecnicas)
    qualidade = pd.DataFrame({
        'Indicador': ['Linhas processadas', 'Matches SalesForce', 'Sem match SalesForce',
                      'Data mínima', 'Data máxima', 'Conflitos/duplicidades Salesforce'],
        'Valor': [len(df), int((df['Match SalesForce'] == 'Sim').sum()),
                  int((df['Match SalesForce'] == 'Não').sum()),
                  df['Dt. Criação Sol.'].min(), df['Dt. Criação Sol.'].max(), len(conflitos)]})
    dicionario = pd.DataFrame([
        ['Frete Comercial', 'Valor importado do SalesForce pela chave de merge.'],
        ['Chapa', 'Valor importado do SalesForce pela chave de merge.'],
        ['Qtd. Aprovadas Anteriores 30d', 'Cotações Banda/PrecoFixo aprovadas anteriormente, na janela móvel de 30 dias.'],
        ['Soma Var. Aprovada Anterior 30d', 'Soma das variações das aprovações anteriores na mesma janela.'],
        ['Match SalesForce', 'Indica se Frete ou Chapa foi localizado no SalesForce.']],
        columns=['Campo', 'Definição'])

    with pd.ExcelWriter(ARQ_SAIDA, engine='openpyxl') as writer:
        base_saida.to_excel(writer, sheet_name='Base_Processada', index=False)
        for nome, tab in dashboards.items():
            tab.to_excel(writer, sheet_name=nome[:31], index=False)
        qualidade.to_excel(writer, sheet_name='Qualidade', index=False)
        conflitos.to_excel(writer, sheet_name='Conflitos_SF', index=False)
        dicionario.to_excel(writer, sheet_name='Dicionario', index=False)

    wb = load_workbook(ARQ_SAIDA)
    azul = '1F4E78'
    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = PatternFill('solid', fgColor=azul)
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in range(1, ws.max_column + 1):
            vals = [str(ws.cell(r, col).value or '') for r in range(1, min(ws.max_row, 200) + 1)]
            ws.column_dimensions[get_column_letter(col)].width = min(max(max(map(len, vals)) + 2, 12), 38)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00;[Red](#,##0.00);-'
                elif hasattr(cell.value, 'year'):
                    cell.number_format = 'dd/mm/yyyy'

    for nome, coluna in [('Resumo_Cliente_Condicao', 3), ('Resumo_Frete_CIF', 4)]:
        ws = wb[nome]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, coluna).number_format = '0.0%'

    ws = wb['Top_10_Clientes']
    graf = BarChart(); graf.title = 'Top 10 clientes por desconto acumulado'; graf.y_axis.title = 'Desconto'
    graf.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
    graf.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)); graf.height = 7; graf.width = 13
    ws.add_chart(graf, 'D2')

    ws = wb['Resumo_Cliente_Condicao']
    pizza = PieChart(); pizza.title = 'Participação por Cliente Condição'
    pizza.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
    pizza.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)); pizza.height = 7; pizza.width = 10
    ws.add_chart(pizza, 'F2')
    wb.save(ARQ_SAIDA)


def main() -> None:
    cot_raw, sf_raw = ler_bases()
    cot = preparar_cotacoes(cot_raw)
    mapa_sf, conflitos = preparar_salesforce(sf_raw)
    consolidada = adicionar_salesforce(cot, mapa_sf)
    consolidada = calcular_historico_30d(consolidada)
    dashboards = criar_tabelas_dashboard(consolidada)
    exportar_excel(consolidada, dashboards, conflitos)
    print(f'Arquivo criado com sucesso em: {ARQ_SAIDA}')


if __name__ == '__main__':
    main()
